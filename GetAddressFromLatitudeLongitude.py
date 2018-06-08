from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options
from selenium import webdriver
import pdb,time
wbook = load_workbook(filename = 'San Diego, CA.xlsx')
wsheet = wbook['San Diego 2017 Batch 2 Raw Data']
#sheet_ranges['F5'] = 3.14
#wb.save('San Diego, CA.xlsx')
#print(sheet_ranges['F5'].value)
options = Options()
options.add_argument("--headless")
#wbook = load_workbook(filename = 'test.xlsx')
#wsheet = wbook['Sheet1']
row = 100

while (row < 1100):
    try:
        row+=1
        lat = None
        long = None
        contents = None
        driver = webdriver.Firefox(executable_path="geckodriver.exe" , firefox_options=options)
        driver.get('https://www.latlong.net/Show-Latitude-Longitude.html')
        lat = str(wsheet['A{}'.format(row)].value)
        long = str(wsheet['B{}'.format(row)].value)
        driver.find_element_by_xpath("//input[@placeholder='lat']").send_keys(lat)
        driver.find_element_by_xpath("//input[@placeholder='long']").send_keys(long)
        driver.find_element_by_xpath("//button[@title='Show Lat Long converted address on Map']").click()
        time.sleep(2)
        contents = driver.find_element_by_xpath('//*[@id="address"]').get_attribute("value")
        wsheet['E{}'.format(row)] = contents
        print(row, lat, long, contents)
        driver.close()
        if lat is None:
            break
        #print (row,end='.')
    except Exception as e:
        print("$"*50,e)
        time.sleep(10)

wbook.save('San Diego, CA.xlsx')