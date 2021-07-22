from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
import pdb,time
fileName = ''
wbook = load_workbook(filename = fileName)
wsheet = wbook['San Diego 2017 Batch 2 Raw Data']
#sheet_ranges['F5'] = 3.14
#wb.save('San Diego, CA.xlsx')
#print(sheet_ranges['F5'].value)
options = Options()
options.add_argument("--headless")
#wbook = load_workbook(filename = 'test.xlsx')
#wsheet = wbook['Sheet1']
#wsheet['E{}'.format(6)] = '4210-4224 Governor Dr, San Diego, CA 92122, USA'
#wbook.save('San Diego, CA.xlsx')
#pdb.set_trace()
time_start = time.clock()
driver = webdriver.Firefox(executable_path="geckodriver.exe", firefox_options=options)
driver.get('https://www.latlong.net/Show-Latitude-Longitude.html')
row = 5000
err = []
while (row <5680):
    try:
        row+=1
        lat = None
        long = None
        contents = None
        lat = str(wsheet['A{}'.format(row)].value)
        long = str(wsheet['B{}'.format(row)].value)
        driver.find_element_by_xpath("//input[@placeholder='lat']").send_keys(lat)
        driver.find_element_by_xpath("//input[@placeholder='long']").send_keys(long)
        driver.find_element_by_xpath("//button[@title='Show Lat Long converted address on Map']").click()
        contents = driver.find_element_by_xpath('//*[@id="address"]').get_attribute("value")
        while not contents:
            print ('.',end='',flush=True)
            time.sleep(1)
            contents = driver.find_element_by_xpath('//*[@id="address"]').get_attribute("value")
        wsheet['E{}'.format(row)] = contents
        driver.refresh()
        print(row, lat, long, contents)
        if lat is None:
            break

    except Exception as e:
        err.append(row)
        print("$"*50,e)
        row-=1
        #continue
    finally:
        if row%10 == 0:
            print("Saved")
            wbook.save(fileName)
wbook.save(fileName)
print ('*'*100)
time_elapsed = (time.clock() - time_start)
print ("Time Taken in minutes",time_elapsed/60)
print(err)
driver.close()